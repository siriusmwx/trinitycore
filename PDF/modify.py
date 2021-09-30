import time
import json
import pandas
import requests
import openpyxl


class ModifyDataDes:
    def __init__(self, username, passwd):
        self.login_url = "http://driveinsight.ias.huawei.com/iam/v1/user/login"
        self.header = {
            "Accept-Language": "zh-CN",
            "Content-Type": "application/json"
        }
        self.session = requests.Session()
        self.user = username
        self.pwd = passwd
        self.login()

    def login(self):
        """登录DI"""
        data = {"userName": self.user, "passWord": self.pwd}
        response = self.session.post(self.login_url,
                                     json=data,
                                     headers=self.header)
        assert response.status_code == 200, "DI登陆失败"
        res_json = json.loads(response.text)
        # print("登陆信息：{}".format(res_json))
        print("登录DI成功")
        # 更新header信息
        self.header["token"] = res_json.get("message").get("token")
        self.header["userId"] = str(res_json.get("message").get("userId"))
        self.header["userName"] = res_json.get("message").get("userName")
        self.header["projectId"] = str(
            res_json.get("message").get("itemInfo").get(
                "projectId"))  # 需要带 否则会报token错误

    def modifydatades(self, destIndex, indexId, addSegmDescription):
        url = "http://driveinsight.ias.huawei.com/manor/v1/update/segments/designated"
        data = {
            "destIndex":
            destIndex,
            "updContents": [{
                "indexId": indexId,
                "addSegmDescription": addSegmDescription
            }]
        }

        response = self.session.post(url,
                                     headers=self.header,
                                     data=json.dumps(data))
        response = json.loads(response.text)

        return response

    def GetDataIndexId(self, dataname):
        name = dataname

        datetemp = name.split("_")[1].split("-")[0]
        date = datetemp.split("/")[0] + "-" + datetemp.split(
            "/")[1] + "-" + datetemp.split("/")[2]
        StrArrtemp_start = date + " " + datetemp.split("/")[3]
        StrArrtemp_end = date + " " + name.split("-")[1]

        # 转换成时间数组
        timeArray = time.strptime(StrArrtemp_start, "%Y-%m-%d %H:%M:%S")
        # 转换成时间戳
        starttime = str(int(time.mktime(timeArray) * 1000000))

        # 转换成时间数组
        timeArray = time.strptime(StrArrtemp_end, "%Y-%m-%d %H:%M:%S")
        # 转换成时间戳
        endtime = str(int(time.mktime(timeArray) * 1000000))

        IndexId = name.split("_")[0] + "_" + starttime + "_" + endtime
        return IndexId


"""305_2021/06/08/15:07:56-15:08:15	11111
"""
if __name__ == '__main__':
    test = ModifyDataDes('user', 'password')
    sheet_name = "Sheet1"
    datas = pandas.read_excel("example.xlsx", sheet_name=sheet_name).values
    workbook = openpyxl.load_workbook("example.xlsx")
    sheet = workbook[sheet_name]

    for i in range(len(datas)):
        IndexId = test.GetDataIndexId(datas[i][0])
        res = test.modifydatades("segmentsSceneModel", IndexId, datas[i][1])
        try:
            if res["message"] == "success":
                print('%s------>%s------>%s' %
                      (datas[i][0], datas[i][1], res["message"]))
                sheet.cell(row=i + 2, column=3, value='pass')
            else:
                print('%s------>%s------>%s' %
                      (datas[i][0], datas[i][1], "failed"))
                sheet.cell(row=i + 2, column=3, value='fail')
        except KeyError:
            continue

    workbook.save("example.xlsx")
